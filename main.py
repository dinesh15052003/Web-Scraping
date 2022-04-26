from openpyxl import load_workbook        #for working with excel
from selenium import webdriver            #for automating webpage
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
import os        #for interacting with operating system
import shutil    #for automating operations on files
import time

workbook = load_workbook("DATA.xlsx")
spreadsheet = workbook.active

download_path = "D:\Triple Threat\Output"
options = Options()
options.add_experimental_option('prefs', {"download.default_directory": download_path,
    "download.prompt_for_download": False, "download.directory_upgrade": True,
    "plugins.plugins_disabled": ["Chrome PDF Viewer"], "plugins.always_open_pdf_externally": True})
d = 1
for m in range(2, spreadsheet.max_row+1):
    if spreadsheet.cell(m, 1).value != None:
        driver = webdriver.Chrome("chromedriver_win32/chromedriver.exe", options=options)
        driver.maximize_window()
        driver.get("https://enrollmentdemo.solartis.net/Quote.xhtml")

        driver.find_element(By.XPATH,'''//*[@id="quoteSummaryDiv"]/table/tbody/tr/td/table/tbody/tr/td/a/img''').click()
        driver.find_element(By.ID,"EAANumber").send_keys(spreadsheet.cell(m, 1).value)
        driver.find_element(By.ID,"firstName").send_keys(spreadsheet.cell(m, 2).value)
        if spreadsheet.cell(m, 3).value!=None:
            driver.find_element(By.ID,"middleName").send_keys(spreadsheet.cell(m, 3).value)
        driver.find_element(By.ID,"lastName").send_keys(spreadsheet.cell(m, 4).value)
        driver.find_element(By.ID,"insuredDateOfBirth").send_keys(spreadsheet.cell(m, 5).value)
        driver.find_element(By.ID,"address1").send_keys(spreadsheet.cell(m, 6).value)
        if spreadsheet.cell(m, 7).value != None:
            driver.find_element(By.ID,"address2").send_keys(spreadsheet.cell(m, 7).value)
        driver.find_element(By.ID,"city").send_keys(spreadsheet.cell(m, 8).value)
        Select(driver.find_element(By.ID,"state")).select_by_visible_text(spreadsheet.cell(m, 9).value)
        driver.find_element(By.CLASS_NAME,"rf-au-fnt.rf-au-inp").send_keys(spreadsheet.cell(m, 10).value)
        driver.find_element(By.ID,"email").send_keys(spreadsheet.cell(m, 11).value)
        driver.find_element(By.ID,"phone").send_keys(spreadsheet.cell(m, 12).value)

        WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "planNameInput"))).click()

        if (spreadsheet.cell(m, 13).value == "Member Only Plan"):
            driver.find_element(By.ID,"planNameItem0").click()
        elif (spreadsheet.cell(m, 13).value == "Member and Spouse"):
            driver.find_element(By.ID,"planNameItem1").click()
        elif (spreadsheet.cell(m, 13).value == "Member and Children"):
            driver.find_element(By.ID,"planNameItem2").click()
        elif (spreadsheet.cell(m, 13).value == "Member and Spouse and Children"):
            driver.find_element(By.ID,"planNameItem3").click()

        WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "coverageLimitInput"))).click()
        #driver.find_element(By.ID, "coverageLimitInput").click()
        if spreadsheet.cell(m, 14).value == "$50,000":
            driver.find_element(By.ID, "coverageLimitItem0").click()
        elif spreadsheet.cell(m, 14).value == "$75,000":
            driver.find_element(By.ID, "coverageLimitItem1").click()
        elif spreadsheet.cell(m, 14).value == "$100,000":
            driver.find_element(By.ID, "coverageLimitItem2").click()
        elif spreadsheet.cell(m, 14).value == "$150,000":
            driver.find_element(By.ID, "coverageLimitItem3").click()
        elif spreadsheet.cell(m, 14).value == "$200,000":
            driver.find_element(By.ID, "coverageLimitItem4").click()
        elif spreadsheet.cell(m, 14).value == "$250,000":
            driver.find_element(By.ID, "coverageLimitItem5").click()

        driver.find_element(By.ID, "submitCertReq2").click()

        for i in range(int(spreadsheet.cell(m, 15).value)):
            WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "j_idt53"))).click()
            #driver.find_element_by_id("j_idt53").click()  # beneficiary
            WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "beneficiaryFirstName"))).click()
            driver.find_element(By.ID,"beneficiaryFirstName").send_keys(spreadsheet.cell(m + i, 16).value)
            if spreadsheet.cell(m + i, 17).value != None:
                driver.find_element(By.ID,"beneficiaryMiddleName").click()
                driver.find_element(By.ID,"beneficiaryMiddleName").send_keys(spreadsheet.cell(m + i, 17).value)
            driver.find_element(By.ID,"beneficiaryLastName").click()
            driver.find_element(By.ID,"beneficiaryLastName").send_keys(spreadsheet.cell(m + i, 18).value)
            driver.find_element(By.ID,"relationship").click()
            driver.find_element(By.ID,"relationship").send_keys(spreadsheet.cell(m + i, 19).value)
            driver.find_element(By.ID,"Beneficiarypercentage").clear()
            driver.find_element(By.ID,"Beneficiarypercentage").send_keys(spreadsheet.cell(m + i, 20).value)
            driver.find_element(By.ID,"j_idt127").click()  # save

        WebDriverWait(driver, 200).until(EC.presence_of_element_located((By.ID, "j_idt55"))).click()
        #driver.find_element_by_id("j_idt55").click()  # continue

        for i in range(int(spreadsheet.cell(m, 21).value)):
            WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "j_idt62"))).click()
            #driver.find_element_by_id("j_idt62").click()  # Add dependant
            WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "dependentFirstName"))).click()
            driver.find_element(By.ID,"dependentFirstName").send_keys(spreadsheet.cell(m + i, 22).value)
            if (spreadsheet.cell(m + i, 23).value != None):
                driver.find_element(By.ID,"middleInitial").click()
                driver.find_element(By.ID,"middleInitial").send_keys(spreadsheet.cell(m + i, 23).value)
            driver.find_element(By.ID,"dependentLastName").click()
            driver.find_element(By.ID,"dependentLastName").send_keys(spreadsheet.cell(m + i, 24).value)
            driver.find_element(By.ID,"dependentDateOfBirth").click()
            driver.find_element(By.ID,"dependentDateOfBirth").send_keys(spreadsheet.cell(m + i, 25).value)
            if (spreadsheet.cell(m + i, 26).value == "Spouse"):
                driver.find_element(By.ID,"relationship:0").click()
            elif (spreadsheet.cell(m + i, 26).value == "Child"):
                driver.find_element(By.ID,"relationship:1").click()
            if (spreadsheet.cell(m + i, 27).value == "Male"):
                driver.find_element(By.ID,"gender:0").click()
            elif (spreadsheet.cell(m + i, 27).value == "Female"):
                driver.find_element(By.ID,"gender:1").click()
            driver.find_element(By.ID,"j_idt174").click()  # Save

        if spreadsheet.cell(m, 21).value > 0:
            WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "j_idt64"))).click()
            #driver.find_element_by_id("j_idt64").click()  # continue

        WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, '''//*[@id="cardType"]/option[3]'''))).click()
        #driver.find_element_by_xpath('''//*[@id="cardType"]/option[3]''').click()
        driver.find_element(By.ID,"cardNumber").send_keys("5555555555554444")
        driver.find_element(By.ID,"cvvNumber").send_keys('123')
        driver.find_element(By.ID,"confirmEmailAddress").send_keys(spreadsheet.cell(m, 11).value)
        driver.find_element(By.ID,"j_idt109").click()  # continue
        WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "agree"))).click()
        #driver.find_element_by_id("agree").click()  # checkbox
        WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "pay"))).click()
        #driver.find_element_by_id("pay").click()

        WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "ReceiptPDF"))).click()
        #driver.find_element_by_id("ReceiptPDF").click()
        time.sleep(15)
        filename = max([download_path + "\\" + f for f in os.listdir(download_path)], key=os.path.getctime)
        shutil.move(filename, os.path.join(download_path, r"TestData"+str(d)+"receipt.pdf"))

        driver.find_element(By.ID, "COIPDF").click()
        time.sleep(15)
        filename = max([download_path + "\\" + f for f in os.listdir(download_path)], key=os.path.getctime)
        shutil.move(filename, os.path.join(download_path, r"TestData"+str(d)+"COI.pdf"))

        d = d + 1
        driver.close()