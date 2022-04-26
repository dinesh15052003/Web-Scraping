from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

workbook = load_workbook("Test Data.xlsx")
spreadsheet = workbook.active

driver = webdriver.Chrome("chromedriver_win32/chromedriver.exe")
driver.maximize_window()
driver.get("https://enrollmentdemo.solartis.net/Quote.xhtml")

driver.find_element(By.XPATH,'''//*[@id="quoteSummaryDiv"]/table/tbody/tr/td/table/tbody/tr/td/a/img''').click()
m=2
driver.find_element(By.ID,"EAANumber").send_keys(spreadsheet.cell(m, 1).value)
driver.find_element(By.ID,"firstName").send_keys(spreadsheet.cell(m, 2).value)
if spreadsheet.cell(m, 3).value!=None:
    driver.find_element(By.ID,"middleName").send_keys(spreadsheet.cell(m, 3).value)
driver.find_element(By.ID,"lastName").send_keys(spreadsheet.cell(m, 4).value)
driver.find_element(By.ID,"insuredDateOfBirth").send_keys(spreadsheet.cell(m, 5).value)
driver.find_element(By.ID,"address1").send_keys(spreadsheet.cell(m, 6).value)
if spreadsheet.cell(m, 7).value != None:
    driver.find_element(By.ID,"address2").send_keys(spreadsheet.cell(m, 7).value)
driver.find_element(By.ID,"city").send_keys(spreadsheet.cell(m, 8).value)
Select(driver.find_element(By.ID,"state")).select_by_visible_text(spreadsheet.cell(m, 9).value)
driver.find_element(By.CLASS_NAME,"rf-au-fnt.rf-au-inp").send_keys(spreadsheet.cell(m, 10).value)
driver.find_element(By.ID,"email").send_keys(spreadsheet.cell(m, 11).value)
driver.find_element(By.ID,"phone").send_keys(spreadsheet.cell(m, 12).value)

WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "planNameInput"))).click()

if (spreadsheet.cell(m, 13).value == "Member Only Plan"):
    driver.find_element(By.ID,"planNameItem0").click()
elif (spreadsheet.cell(m, 13).value == "Member and Spouse"):
    driver.find_element(By.ID,"planNameItem1").click()
elif (spreadsheet.cell(m, 13).value == "Member and Children"):
    driver.find_element(By.ID,"planNameItem2").click()
elif (spreadsheet.cell(m, 13).value == "Member and Spouse and Children"):
    driver.find_element(By.ID,"planNameItem3").click()

WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "coverageLimitInput"))).click()
#driver.find_element(By.ID, "coverageLimitInput").click()
if spreadsheet.cell(m, 14).value == "$50,000":
    driver.find_element(By.ID, "coverageLimitItem0").click()
elif spreadsheet.cell(m, 14).value == "$75,000":
    driver.find_element(By.ID, "coverageLimitItem1").click()
elif spreadsheet.cell(m, 14).value == "$100,000":
    driver.find_element(By.ID, "coverageLimitItem2").click()
elif spreadsheet.cell(m, 14).value == "$150,000":
    driver.find_element(By.ID, "coverageLimitItem3").click()
elif spreadsheet.cell(m, 14).value == "$200,000":
    driver.find_element(By.ID, "coverageLimitItem4").click()
elif spreadsheet.cell(m, 14).value == "$250,000":
    driver.find_element(By.ID, "coverageLimitItem5").click()

driver.find_element(By.ID, "submitCertReq2").click()

for i in range(int(spreadsheet.cell(m, 15).value)):
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "j_idt53"))).click()
    # driver.find_element_by_id("j_idt53").click()  # beneficiary
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "beneficiaryFirstName"))).click()
    driver.find_element(By.ID, "beneficiaryFirstName").send_keys(spreadsheet.cell(m + i, 16).value)
    if spreadsheet.cell(m + i, 17).value != None:
        driver.find_element(By.ID, "beneficiaryMiddleName").click()
        driver.find_element(By.ID, "beneficiaryMiddleName").send_keys(spreadsheet.cell(m + i, 17).value)
    driver.find_element(By.ID, "beneficiaryLastName").click()
    driver.find_element(By.ID, "beneficiaryLastName").send_keys(spreadsheet.cell(m + i, 18).value)
    driver.find_element(By.ID, "relationship").click()
    driver.find_element(By.ID, "relationship").send_keys(spreadsheet.cell(m + i, 19).value)
    driver.find_element(By.ID, "Beneficiarypercentage").clear()
    driver.find_element(By.ID, "Beneficiarypercentage").send_keys(spreadsheet.cell(m + i, 20).value)
    driver.find_element(By.ID, "j_idt127").click()  # save

WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "j_idt55"))).click()

for i in range(int(spreadsheet.cell(m, 21).value)):
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "j_idt62"))).click()
    # driver.find_element_by_id("j_idt62").click()  # Add dependant
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "dependentFirstName"))).click()
    driver.find_element(By.ID, "dependentFirstName").send_keys(spreadsheet.cell(m + i, 22).value)
    if (spreadsheet.cell(m + i, 23).value != None):
        driver.find_element(By.ID, "middleInitial").click()
        driver.find_element(By.ID, "middleInitial").send_keys(spreadsheet.cell(m + i, 23).value)
    driver.find_element(By.ID, "dependentLastName").click()
    driver.find_element(By.ID, "dependentLastName").send_keys(spreadsheet.cell(m + i, 24).value)
    driver.find_element(By.ID, "dependentDateOfBirth").click()
    driver.find_element(By.ID, "dependentDateOfBirth").send_keys(spreadsheet.cell(m + i, 25).value)
    if (spreadsheet.cell(m + i, 26).value == "Spouse"):
        driver.find_element(By.ID, "relationship:0").click()
    elif (spreadsheet.cell(m + i, 26).value == "Child"):
        driver.find_element(By.ID, "relationship:1").click()
    if (spreadsheet.cell(m + i, 27).value == "Male"):
        driver.find_element(By.ID, "gender:0").click()
    elif (spreadsheet.cell(m + i, 27).value == "Female"):
        driver.find_element(By.ID, "gender:1").click()
    driver.find_element(By.ID, "j_idt174").click()  # Save

if spreadsheet.cell(m, 21).value > 0:
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "j_idt64"))).click()
    # driver.find_element_by_id("j_idt64").click()  # continue

WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, '''//*[@id="cardType"]/option[3]'''))).click()
# driver.find_element_by_xpath('''//*[@id="cardType"]/option[3]''').click()
driver.find_element(By.ID, "cardNumber").send_keys("5555555555554444")
driver.find_element(By.ID, "cvvNumber").send_keys('123')
driver.find_element(By.ID, "confirmEmailAddress").send_keys(spreadsheet.cell(m, 11).value)
driver.find_element(By.ID, "j_idt109").click()  # continue
#
# WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "agree"))).click()
# #driver.find_element_by_id("agree").click()  # checkbox
# WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "pay"))).click()
# #driver.find_element_by_id("pay").click()